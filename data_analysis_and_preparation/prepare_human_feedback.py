import re, random, os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_data_from_excel(data_wb, key_wb):
    """
    data_wb: filepath to excel file with the evaluations, as received from the evaluators
    key_wv: filepath to excel file that contains the letter ids and model names"""
    wb = load_workbook(data_wb)
    wb_key = load_workbook(key_wb)
    ws = wb["questionnaire"]  
    ws_key = wb_key["questionnaire"]
    data = {
        "letter_id": [],
        "n_footnote": [],
        "model": [],
        "text_footnote": [],
        "style": [],
        "usefullness": [],
        "correctness": [],
        "fact_check": [],
        "commentary": []
    }

    for row_num in range(2, ws.max_row+1):
        id = ws_key.cell(row=row_num, column=1).value
        if id:
            id_split = id.split("_")
            continue
        else:
            data["letter_id"].append(id_split[0])
            data["n_footnote"].append(id_split[1])
            data["model"].append(ws_key.cell(row=row_num, column=2).value)
            data["text_footnote"].append(ws.cell(row=row_num, column=2).value)
            data["usefullness"].append(ws.cell(row=row_num, column=4).value)
            data["style"].append(ws.cell(row=row_num, column=3).value)
            data["correctness"].append(ws.cell(row=row_num, column=5).value)
            data["fact_check"].append(ws.cell(row=row_num, column=6).value)
            data["commentary"].append(ws.cell(row=row_num, column=7).value)
    
    return pd.DataFrame(data)


def remove_other_fns(n_footnote, text):
    regex = rf"__(?!{n_footnote}\b)\w+"  # (?!30\b): Negative lookahead that ensures the footnote in question does not immediately follow __.
    return re.sub(regex, "", text)

def fill_excel_with_data(df, infile_path, outfile_path):
    """
    This function takes a pandas DataFrame, reads an existing Excel file, 
    and writes data to it as specified.
    
    Parameters:
        df (pd.DataFrame): DataFrame containing data to write.
        infile_path (str): Path to the existing Excel file to read.
        outfile_path (str): Path to save the modified Excel file.
    """
    # Load the existing workbook 
    wb = load_workbook(infile_path)
    ws = wb["questionnaire"]  

    # Define the color fill
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # more of an orange
    blue_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Start inserting data from row 2 (headers are in row 1)
    start_row = 2

    for index, row in df.iterrows():
        row_num = start_row + index*5  # every letter takes 5 rows

        # fill in id, combination of letter_id and n_footnote in first column
        ws.cell(row=row_num, column=1, value=f"{row.letter_id}_{row.n_footnote}")


        # Step 1: Set title and footnote number in first row of the letter
        
        ws.cell(row=row_num, column=3, value=row['title_letter']) # Column A
        ws.cell(row=row_num, column=4, value=f'Fussnote {row["n_footnote"]}')

        # step 1.1 coloring the cells
        # yellow row
        for col in range(3, ws.max_column + 1):
            ws.cell(row=row_num, column=col).fill = yellow_fill
        
        # set remaining cells to alternating shades of blue
        for i in range(1,5):
            for col in range(4, ws.max_column + 1):
                if i%2:
                    ws.cell(row=row_num+i, column=col).fill = light_blue_fill
                else:
                    ws.cell(row=row_num+i, column=col).fill = blue_fill


    # Step 2: write the letter text and merge cells
        sentence = remove_other_fns(row["n_footnote"], row["text_sentence"])
        ws.cell(row=row_num + 1, column=3, value=sentence)
        ws.merge_cells(start_row=row_num + 1, start_column=3, end_row=row_num + 4, end_column=3)

        

        # Step 3: Insert the models (also the human footnote) into the df
        # do it in random order, save the model id in second column
        model_names = df.columns[4:8]  # places of the models in the df
        rnd_placements = random.sample([1,2,3,4], 4)
        for rnd_placement, model_name in zip(rnd_placements, model_names):
            ws.cell(row=row_num + rnd_placement, column=2, value=model_name)
            ws.cell(row=row_num + rnd_placement, column=4, value=row[model_name])

    # Save the modified workbook
    wb.save(outfile_path)

def sample_from_test_set(footnote_df):

    # get letter ids and n-footnotes from the testset
    test_fns = [filename.split(".")[0] for filename in os.listdir("../../data/prompts/instruct_add/test")]

    filtered_rows = []
    for test_fn in test_fns:
        letter_id = int(test_fn.split("_")[0])
        n_footnote = int(test_fn.split("_")[1])
        filtered_rows.append(footnote_df[(footnote_df['letter_id'] == letter_id) & (footnote_df['n_footnote'] == n_footnote)])

    df = pd.concat(filtered_rows, ignore_index=True)

    # remove the double label columns for stratification
    df = df[~df["label"].str.contains(",")]

    # remove labels that are only in there bc of the new classification
    df = df[~df["label"].isin(["missing", "lex", "lex_dict"])]


    # add a column to the df with the language of the sentence
    # should be able to do it from the xml_sent column, via a regex matching the xml:lang attribute
    lang_regex = r"xml:lang=\"(\w+)\""
    langs = []
    for _, row in df.iterrows():
        lang = re.search(lang_regex, row.xml_sentence).group(1)
        langs.append(lang)

    df["lang"] = langs

    strat_sample_de = stratify_sample_by_label(df[df["lang"]=="de"]) 
    strat_sample_la = stratify_sample_by_label(df[df["lang"]=="la"])



    return pd.concat([strat_sample_de, strat_sample_la], ignore_index=True)

def stratify_sample_by_label(df, n=20):
    sample = df.groupby('label', group_keys=False).apply(
        lambda x: x.sample(frac=n / len(df), random_state=42)
    ).reset_index(drop=True)

    if sample.shape[0] < n:  # Check if the sample size is less than n
        remaining_indices = df.index.difference(sample.index)  # Indices not in the sample
        additional_sample = df.loc[remaining_indices].sample(
            n=n - sample.shape[0], random_state=42, replace=False
        )
        sample = pd.concat([sample, additional_sample], ignore_index=True)

    return sample

def input_letter_head(sample_df, n_column):
    """find the letter titles and put them into the df"""
    titles = []
    letters_dir = "../../data/downsized_letters"
    title_regex = r"<titleStmt>[^<]*<title[^>]+>([^<]+)"

    for _, row in sample_df.iterrows():
        with open(os.path.join(letters_dir, f"{row.letter_id}.xml"), "r", encoding="utf-8") as infile:
            letter_text = infile.read()
        title_match = re.search(title_regex, letter_text)
        if title_match:
            titles.append(title_match.group(1))
        else:
            print(f"No title in {row.letter_id}")
            titles.append("NaN")
    sample_df.insert(n_column, "title_letter", titles)
        